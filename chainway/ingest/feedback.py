"""★ 業務／門市市場調查回饋匯入 —— 你要「自行填入暢銷/滯銷理由」的那一塊。

三種填寫方式，都指向同一份資料：
  A. 直接編輯 data/feedback/sales_feedback.csv（Excel 開得起來，最快）
  B. 用 `python -m chainway.cli feedback template` 產生附下拉選單的 xlsx 發給業務填
  C. 在網頁後台「回饋登錄」頁逐筆填寫（會寫回同一個 CSV）

同一個貨號可以有多筆回饋（不同門市、不同時間、不同來源）—— 這是刻意的設計，
因為「台北說太貴、台中說很好賣」本身就是重要訊號，不該被覆蓋掉。
"""

from __future__ import annotations

import datetime as dt
from pathlib import Path
from typing import Any

import pandas as pd

from ..config import Config, get_config

FEEDBACK_FILE = "sales_feedback.csv"

COLUMNS = [
    "sku", "style_code", "season", "category", "verdict", "reason_tags",
    "reason_text", "source", "respondent", "store_or_region", "survey_date",
    "confidence", "suggested_action", "follow_up_note",
]

CONFIDENCE_LEVELS = ["HIGH", "MEDIUM", "LOW"]


def feedback_path(cfg: Config | None = None) -> Path:
    cfg = cfg or get_config()
    return cfg.path("feedback") / FEEDBACK_FILE


def ensure_feedback_file(cfg: Config | None = None) -> Path:
    """檔案不存在就用 TEMPLATE 開一份（保留範例列讓填寫人看得懂格式）。"""
    cfg = cfg or get_config()
    path = feedback_path(cfg)
    if path.exists():
        return path
    template = cfg.path("feedback") / "sales_feedback_TEMPLATE.csv"
    if template.exists():
        path.write_text(template.read_text(encoding="utf-8"), encoding="utf-8")
    else:
        pd.DataFrame(columns=COLUMNS).to_csv(path, index=False, encoding="utf-8-sig")
    return path


def load_feedback(cfg: Config | None = None) -> pd.DataFrame:
    cfg = cfg or get_config()
    path = ensure_feedback_file(cfg)
    df = pd.read_csv(path, dtype=str, keep_default_na=False)
    for col in COLUMNS:
        if col not in df.columns:
            df[col] = ""
    df = df[COLUMNS].copy()
    df["sku"] = df["sku"].str.strip()
    df = df[df["sku"].ne("")]
    df["survey_date"] = pd.to_datetime(df["survey_date"], errors="coerce")
    df["reason_tag_list"] = df["reason_tags"].fillna("").apply(
        lambda s: [t.strip().upper() for t in str(s).replace(",", "|").split("|") if t.strip()]
    )
    return df.reset_index(drop=True)


def append_feedback(record: dict[str, Any], cfg: Config | None = None) -> Path:
    """新增一筆回饋（網頁後台與 CLI 都走這個函式）。"""
    cfg = cfg or get_config()
    path = ensure_feedback_file(cfg)
    row = {col: str(record.get(col, "") or "") for col in COLUMNS}
    if not row["sku"]:
        raise ValueError("sku（貨號）為必填")
    if not row["survey_date"]:
        row["survey_date"] = dt.date.today().isoformat()
    if isinstance(record.get("reason_tags"), (list, tuple)):
        row["reason_tags"] = "|".join(record["reason_tags"])

    existing = pd.read_csv(path, dtype=str, keep_default_na=False) if path.stat().st_size else pd.DataFrame(columns=COLUMNS)
    for col in COLUMNS:
        if col not in existing.columns:
            existing[col] = ""
    out = pd.concat([existing[COLUMNS], pd.DataFrame([row])], ignore_index=True)
    out.to_csv(path, index=False, encoding="utf-8-sig")
    return path


def validate_feedback(df: pd.DataFrame, cfg: Config | None = None) -> pd.DataFrame:
    """檢查填寫品質：不合法的 verdict / 不存在的標籤 / 缺文字說明。"""
    cfg = cfg or get_config()
    valid_verdicts = {v["code"] for v in cfg.feedback_tags.get("verdicts", [])}
    valid_sources = {s["code"] for s in cfg.feedback_tags.get("sources", [])}
    valid_tags = {t["code"] for t in cfg.all_reason_tags()}
    valid_actions = {a["code"] for a in cfg.feedback_tags.get("actions", [])}

    issues: list[dict[str, Any]] = []
    for i, row in df.iterrows():
        problems = []
        if row["verdict"] and row["verdict"].upper() not in valid_verdicts:
            problems.append(f"verdict 值 '{row['verdict']}' 不在允許清單")
        if row["source"] and row["source"].upper() not in valid_sources:
            problems.append(f"source 值 '{row['source']}' 不在允許清單")
        unknown = [t for t in row["reason_tag_list"] if t not in valid_tags]
        if unknown:
            problems.append(f"未知理由標籤：{', '.join(unknown)}")
        if row["suggested_action"] and row["suggested_action"].upper() not in valid_actions:
            problems.append(f"suggested_action '{row['suggested_action']}' 不在允許清單")
        if not row["reason_tag_list"] and not str(row["reason_text"]).strip():
            problems.append("理由標籤與文字說明皆空白，這筆無法用於診斷")
        if problems:
            issues.append({"row": int(i) + 2, "sku": row["sku"], "issues": "；".join(problems)})
    return pd.DataFrame(issues)


def summarize_feedback(df: pd.DataFrame, cfg: Config | None = None) -> pd.DataFrame:
    """把多筆回饋收斂成「一個貨號一列」，供合併主表使用。"""
    cfg = cfg or get_config()
    if df.empty:
        return pd.DataFrame(columns=["sku", "fb_n", "fb_verdict", "fb_tags", "fb_groups", "fb_texts"])

    rows = []
    for sku, g in df.groupby("sku"):
        tags: list[str] = [t for lst in g["reason_tag_list"] for t in lst]
        verdicts = [v.upper() for v in g["verdict"] if v]
        # 多數決；平手時以最新一筆為準
        verdict = ""
        if verdicts:
            counts = pd.Series(verdicts).value_counts()
            top = counts[counts == counts.max()].index.tolist()
            verdict = top[0] if len(top) == 1 else g.sort_values("survey_date")["verdict"].iloc[-1].upper()
        groups = sorted({cfg.reason_tag_group(t) for t in tags})
        texts = [f"[{r['source'] or '?'}/{r['store_or_region'] or '?'}] {r['reason_text']}"
                 for _, r in g.iterrows() if str(r["reason_text"]).strip()]
        rows.append({
            "sku": sku,
            "fb_n": len(g),
            "fb_verdict": verdict,
            "fb_tags": "|".join(sorted(set(tags))),
            "fb_groups": "|".join(groups),
            "fb_texts": " ｜ ".join(texts),
        })
    return pd.DataFrame(rows)


def make_excel_template(out_path: Path | None = None, cfg: Config | None = None) -> Path:
    """產生帶下拉選單的 xlsx，發給業務/門市填寫（不用背代碼）。"""
    cfg = cfg or get_config()
    out_path = out_path or (cfg.path("feedback") / "回饋表_發放用.xlsx")

    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    ws = wb.active
    ws.title = "回饋填寫"

    headers_zh = {
        "sku": "貨號*", "style_code": "款號", "season": "季別", "category": "品類",
        "verdict": "判定*", "reason_tags": "理由代碼(可複選,用|分隔)*", "reason_text": "文字說明*",
        "source": "回饋來源*", "respondent": "填寫人", "store_or_region": "門市/區域",
        "survey_date": "調查日期", "confidence": "把握度", "suggested_action": "建議行動",
        "follow_up_note": "後續備註",
    }
    ws.append([headers_zh[c] for c in COLUMNS])
    ws.freeze_panes = "A2"
    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 34 if col in ("reason_tags", "reason_text") else 16

    # 代碼對照分頁（同時當作下拉選單的來源）
    ref = wb.create_sheet("代碼對照")
    ref.append(["類型", "代碼", "中文", "群組"])
    for v in cfg.feedback_tags.get("verdicts", []):
        ref.append(["判定", v["code"], v["zh"], ""])
    for s in cfg.feedback_tags.get("sources", []):
        ref.append(["來源", s["code"], s["zh"], ""])
    for t in cfg.all_reason_tags():
        ref.append(["理由", t["code"], t["zh"], t["group_zh"]])
    for a in cfg.feedback_tags.get("actions", []):
        ref.append(["行動", a["code"], a["zh"], ""])
    for i, w in enumerate([10, 20, 30, 16], start=1):
        ref.column_dimensions[get_column_letter(i)].width = w

    def add_dv(col_code: str, values: list[str]) -> None:
        idx = COLUMNS.index(col_code) + 1
        letter = get_column_letter(idx)
        dv = DataValidation(type="list", formula1='"' + ",".join(values) + '"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{letter}2:{letter}2000")

    add_dv("verdict", [v["code"] for v in cfg.feedback_tags.get("verdicts", [])])
    add_dv("source", [s["code"] for s in cfg.feedback_tags.get("sources", [])])
    add_dv("confidence", CONFIDENCE_LEVELS)
    add_dv("category", cfg.category_codes)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path
